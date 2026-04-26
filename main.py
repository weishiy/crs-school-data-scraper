import os
import re
from typing import Dict, List
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from pypinyin import lazy_pinyin
from datetime import datetime


BASE_URL = "https://www.crs.jsj.edu.cn"
LIST_URL = "https://www.crs.jsj.edu.cn/aproval/orglists"
REFERER = "https://www.crs.jsj.edu.cn/index/sort/1006"

CRS_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Referer": REFERER,
    "Origin": BASE_URL,
    "Content-Type": "application/x-www-form-urlencoded",
}

EXTERNAL_HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8",
}

# 可直接扩展，例如：COUNTRIES = ["英国", "加拿大"]
COUNTRIES = ["美国"]

# 联系方式自动搜索时，最多处理多少条。None 表示全部处理。
MAX_ENRICH_RECORDS = None

# 对搜索候选页最低分要求。分太低时宁可留空，也不要乱填。
MIN_CONTACT_PAGE_SCORE = 10


University_985 = {
    "北京大学", "中国人民大学", "清华大学", "北京航空航天大学", "北京理工大学",
    "中国农业大学", "北京师范大学", "中央民族大学",
    "南开大学", "天津大学",
    "大连理工大学", "东北大学",
    "吉林大学",
    "哈尔滨工业大学",
    "复旦大学", "同济大学", "上海交通大学", "华东师范大学",
    "南京大学", "东南大学",
    "浙江大学",
    "中国科学技术大学",
    "厦门大学",
    "山东大学", "中国海洋大学",
    "武汉大学", "华中科技大学",
    "中南大学", "湖南大学", "国防科技大学",
    "中山大学", "华南理工大学",
    "四川大学", "电子科技大学",
    "重庆大学",
    "西安交通大学", "西北工业大学",
    "西北农林科技大学",
    "兰州大学",
}

University_211 = {
    "北京大学", "中国人民大学", "清华大学", "北京交通大学", "北京工业大学", "北京航空航天大学",
    "北京理工大学", "北京科技大学", "北京化工大学", "北京邮电大学", "中国农业大学",
    "北京林业大学", "北京中医药大学", "北京师范大学", "北京外国语大学", "中国传媒大学",
    "中央财经大学", "对外经济贸易大学", "中国政法大学", "中央民族大学", "华北电力大学",
    "中国矿业大学", "中国石油大学", "中国地质大学","中国地质大学（北京）",
    "南开大学", "天津大学", "天津医科大学",
    "河北工业大学",
    "太原理工大学",
    "内蒙古大学",
    "辽宁大学", "大连理工大学", "东北大学", "大连海事大学",
    "吉林大学", "延边大学", "东北师范大学",
    "哈尔滨工业大学", "哈尔滨工程大学", "东北农业大学", "东北林业大学",
    "复旦大学", "同济大学", "上海交通大学", "华东理工大学", "东华大学",
    "华东师范大学", "上海外国语大学", "上海财经大学",
    "苏州大学", "南京大学", "东南大学", "南京航空航天大学", "南京理工大学",
    "中国矿业大学", "河海大学", "江南大学", "南京农业大学", "中国药科大学", "南京师范大学",
    "浙江大学",
    "安徽大学", "中国科学技术大学",
    "福州大学", "厦门大学",
    "南昌大学",
    "山东大学", "中国海洋大学", "中国石油大学",
    "郑州大学",
    "武汉大学", "华中科技大学", "中国地质大学", "武汉理工大学", "华中农业大学",
    "华中师范大学", "中南财经政法大学",
    "湖南大学", "中南大学", "湖南师范大学", "国防科技大学",
    "中山大学", "暨南大学", "华南理工大学", "华南师范大学",
    "广西大学",
    "海南大学",
    "四川大学", "西南交通大学", "电子科技大学", "四川农业大学", "西南财经大学",
    "重庆大学", "西南大学",
    "贵州大学",
    "云南大学",
    "西藏大学",
    "西安交通大学", "西北工业大学", "西安电子科技大学", "长安大学",
    "西北大学", "陕西师范大学", "西北农林科技大学",
    "兰州大学",
    "青海大学",
    "宁夏大学",
    "新疆大学", "石河子大学",
}


# ========= 工具函数 =========
def to_pinyin(text: str) -> str:
    return "".join(lazy_pinyin(str(text)))


def clean_text(s: str) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip()


def normalize_foreign_university_name(text: str) -> str:
    text = clean_text(text)
    text = re.sub(r"^(英国|英格兰|苏格兰|威尔士|北爱尔兰|加拿大|美国|澳大利亚|新西兰)", "", text).strip()
    return text


def extract_university_name(full_name: str, category: str) -> str:
    full_name = clean_text(full_name.replace("●", ""))

    if category == "合作办学项目" and "与" in full_name:
        return full_name.split("与")[0].strip()

    school_suffixes = ["职业技术学院", "职业学院", "师范大学", "大学", "学院", "学校"]
    for suffix in school_suffixes:
        idx = full_name.find(suffix)
        if idx != -1:
            return full_name[: idx + len(suffix)].strip()

    return ""


def mark_985_211(university_name: str) -> dict:
    name = university_name.strip()
    is_985 = name in University_985
    is_211 = name in University_211
    return {
        "is_985": is_985,
        "is_211": is_211,
        "is_985_211": is_985 or is_211,
    }


def parse_project_name(project_name: str) -> dict:
    text = clean_text(project_name).replace("●", "")

    result = {
        "china_school": "",
        "foreign_school": "",
        "major": "",
        "level": "",
    }

    if "本科" in text:
        result["level"] = "本科"
    elif "硕士" in text:
        result["level"] = "硕士"
    elif "博士" in text:
        result["level"] = "博士"
    elif "专科" in text:
        result["level"] = "专科"

    match = re.search(r"^(.*?)与(.*?)合作举办", text)
    if match:
        result["china_school"] = clean_text(match.group(1))
        result["foreign_school"] = normalize_foreign_university_name(match.group(2))

    major_patterns = [
        r"合作举办(.*?)专业",
        r"合作举办(.*?)本科教育项目",
        r"合作举办(.*?)硕士学位教育项目",
        r"合作举办(.*?)博士学位教育项目",
    ]
    for pattern in major_patterns:
        m2 = re.search(pattern, text)
        if m2:
            major = clean_text(m2.group(1))
            major = (
                major.replace("本科教育项目", "")
                .replace("硕士学位教育项目", "")
                .replace("博士学位教育项目", "")
            )
            result["major"] = clean_text(major)
            break

    return result


# ========= 一级爬虫：列表页 =========
def fetch_country_html(session: requests.Session, country: str) -> str:
    data = {
        "subjdirect": "",
        "cn_runschool": "",
        "en_runschool": "",
        "local": country,
    }
    response = session.post(LIST_URL, headers=CRS_HEADERS, data=data, timeout=30)
    response.raise_for_status()
    response.encoding = response.apparent_encoding
    return response.text


def extract_records(html: str, country: str) -> List[Dict]:
    soup = BeautifulSoup(html, "html.parser")
    records = []
    current_region = None

    for tr in soup.find_all("tr"):
        tds = tr.find_all("td")
        if not tds:
            continue

        texts = [td.get_text(" ", strip=True) for td in tds]

        if "地区" in texts and "项目/机构" in texts:
            continue

        if len(tds) >= 3:
            region = texts[0]
            category = texts[1]
            name_td = tds[2]
            if not region or not category:
                continue
            current_region = region
        elif len(tds) == 2 and current_region:
            region = current_region
            category = texts[0]
            name_td = tds[1]
            if not category:
                continue
        else:
            continue

        lis = name_td.find_all("li")
        for li in lis:
            full_text = clean_text(li.get_text(" ", strip=True).replace("●", ""))
            if not full_text:
                continue

            link = ""
            for a in li.find_all("a", href=True):
                href = a.get("href")
                if href and "/aproval/detail/" in href:
                    link = urljoin(BASE_URL, href).strip()
                    break

            if not link:
                continue

            university_name = extract_university_name(full_text, category)
            university_tags = mark_985_211(university_name)

            records.append(
                {
                    "country": country,
                    "region": region,
                    "category": category,
                    "university_name": university_name,
                    "name": full_text,
                    "link": link,
                    **university_tags,
                }
            )

    unique_records = []
    seen = set()
    for row in records:
        key = (
            row["country"],
            row["region"],
            row["category"],
            row["university_name"],
            row["name"],
            row["link"],
        )
        if key not in seen:
            seen.add(key)
            unique_records.append(row)

    return unique_records


# ========= 二级爬虫：详情页 =========
def fetch_detail_html(session: requests.Session, detail_url: str) -> str:
    response = session.get(detail_url, headers=CRS_HEADERS, timeout=30)
    response.raise_for_status()
    response.encoding = response.apparent_encoding
    return response.text


def fetch_external_html(session: requests.Session, url: str) -> str:
    response = session.get(url, headers=EXTERNAL_HEADERS, timeout=20)
    response.raise_for_status()
    response.encoding = response.apparent_encoding
    return response.text


def extract_detail_fields(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text("\n", strip=True)

    detail_data = {
        "level": "",
        "duration": "",
        "major_or_course": "",
        "degree_awarded": "",
        "foreign_degree_certificate": "",
        "admission_years": "",
        "admission_end_year": "",
        "is_active": "",
    }

    for tr in soup.find_all("tr"):
        cells = tr.find_all(["td", "th"])
        texts = [cell.get_text(" ", strip=True) for cell in cells if cell.get_text(" ", strip=True)]

        if len(texts) < 2:
            continue

        for i in range(0, len(texts) - 1, 2):
            key = texts[i].strip()
            value = texts[i + 1].strip()

            if key == "办学层次和类别":
                if "硕士" in value:
                    detail_data["level"] = "硕士"
                elif "本科" in value:
                    detail_data["level"] = "本科"
                elif "专科" in value:
                    detail_data["level"] = "专科"
                elif "博士" in value:
                    detail_data["level"] = "博士"
                else:
                    detail_data["level"] = value

            elif key == "学制":
                detail_data["duration"] = value

            elif key == "开设专业或课程":
                detail_data["major_or_course"] = value

            elif key == "颁发证书":
                detail_data["degree_awarded"] = value

                match = re.search(r"外方[：:]\s*(.*?)(?=中方[：:]|$)", value)
                if match:
                    detail_data["foreign_degree_certificate"] = match.group(1).strip()
            elif key == "招生起止年份":
                detail_data["admission_years"] = value

                current_year = datetime.now().year

                if "至今" in value or "长期" in value:
                    detail_data["is_active"] = "是"
                else:
                    match = re.search(r"(\d{4})年[—\-－~～至到](\d{4})年", value)
                    if match:
                        end_year = int(match.group(2))
                        detail_data["admission_end_year"] = end_year

                        if end_year >= current_year:
                            detail_data["is_active"] = "是"
                        else:
                            detail_data["is_active"] = "否（招生年份已过）"
    # 兜底：如果表格里没抓到，再从整个详情页文本里抓
    if not detail_data["foreign_degree_certificate"]:
        full_match = re.search(r"颁发证书[\s\S]*?外方[：:]\s*(.*)", text)
        if full_match:
            detail_data["foreign_degree_certificate"] = full_match.group(1).strip()

    return detail_data

# ========= 自动搜索联系方式页 =========
def build_search_queries(row: dict) -> List[str]:
    project_name = clean_text(row.get("name", ""))
    university = clean_text(row.get("university_name", ""))
    level = clean_text(row.get("level", ""))
    major_or_course = clean_text(row.get("major_or_course", ""))

    parsed = parse_project_name(project_name)
    china_school = parsed["china_school"] or university
    foreign_school = parsed["foreign_school"]
    major = parsed["major"] or major_or_course

    queries = []

    if china_school and major:
        queries.append(f"{china_school} {major} 中外合作办学 招生简章 site:edu.cn")
        queries.append(f"{china_school} {major} 招生 site:edu.cn")
        queries.append(f"{china_school} 国际教育学院 {major} site:edu.cn")
        queries.append(f"{china_school} 国际学院 {major} site:edu.cn")

    if china_school and foreign_school and major:
        queries.append(f"{china_school} {foreign_school} {major} site:edu.cn")

    if china_school and level and major:
        queries.append(f"{china_school} {major} {level} 招生简章 site:edu.cn")

    if project_name:
        short_project = (
            project_name.replace("合作举办", " ")
            .replace("本科教育项目", " ")
            .replace("硕士学位教育项目", " ")
            .replace("博士学位教育项目", " ")
            .replace("专业", " ")
        )
        short_project = clean_text(short_project)
        if china_school and short_project:
            queries.append(f"{china_school} {short_project} site:edu.cn")

    medical_keywords = ["护理", "药学", "临床", "医学", "中医"]
    combined_text = f"{project_name} {major}"
    if china_school and any(k in combined_text for k in medical_keywords):
        queries.append(f"{china_school} 招生办 联系方式 site:edu.cn")
        queries.append(f"{china_school} 本科招生 {major or ''} site:edu.cn")

    seen = set()
    final_queries = []
    for q in queries:
        q = clean_text(q)
        if q and q not in seen:
            seen.add(q)
            final_queries.append(q)

    return final_queries


def score_candidate_page(row: dict, url: str, html: str) -> int:
    soup = BeautifulSoup(html, "html.parser")
    text = clean_text(soup.get_text(" ", strip=True))

    project_name = clean_text(row.get("name", ""))
    university = clean_text(row.get("university_name", ""))
    level = clean_text(row.get("level", ""))
    major_or_course = clean_text(row.get("major_or_course", ""))

    parsed = parse_project_name(project_name)
    china_school = parsed["china_school"] or university
    foreign_school = parsed["foreign_school"]
    major = parsed["major"] or major_or_course

    score = score_candidate_page(row, url, html)
    lower_url = url.lower()

    if china_school and china_school in text:
        score += 4
    if foreign_school and foreign_school in text:
        score += 5
    if major and major in text:
        score += 8
    if level and level in text:
        score += 2
    if "中外合作办学" in text or "合作办学" in text:
        score += 6
    if "招生简章" in text:
        score += 8
    if "招生" in text:
        score += 3
    if "培养模式" in text or any(x in text for x in ["4+0", "2+2", "3+1", "1+3"]):
        score += 2
    if any(x in text for x in ["联系电话", "咨询电话", "电子邮箱", "邮箱"]):
        score += 2

    preferred_parts = [
        "zs", "zsb", "admission", "admissions", "international", "gj", "gjxy",
        "iec", "sie", "yjs", "undergraduate", "recruit", "zhaosheng",
    ]
    if any(x in lower_url for x in preferred_parts):
        score += 3

    bad_parts = ["news", "article", "info", "show", "view", "xw", "xinwen", "notice", "tzgg"]
    if any(x in lower_url for x in bad_parts):
        score -= 2

    if project_name:
        short_name = project_name[:18]
        if short_name in text:
            score += 3

    return score


def search_contact_page_better(row, session):
    from ddgs import DDGS

    university = row.get("university_name", "").strip()
    project_name = row.get("name", "").strip()
    level = row.get("level", "").strip()

    short_project = project_name.replace("合作举办", " ").replace("教育项目", " ").strip()

    queries = [
        f'{university} 招生 {short_project} site:edu.cn',
        f'{university} 招生简章 {short_project} site:edu.cn',
        f'{university} {short_project} 联系方式 site:edu.cn',
        f'{university} 国际学院 {short_project} site:edu.cn',
    ]

    blocked_domains = [
        "zhihu.com", "sohu.com", "163.com", "baidu.com",
        "wenku.baidu.com", "douban.com", "bilibili.com",
        "xiaohongshu.com", "sina.com.cn", "qq.com"
    ]

    candidate_urls = []

    with DDGS() as ddgs:
        for query in queries:
            try:
                results = ddgs.text(query, max_results=5)
            except Exception as e:
                print(f"搜索失败: {query} -> {e}")
                continue

            for r in results:
                url = (r.get("href") or r.get("url") or "").strip()
                if not url:
                    continue
                if any(bad in url for bad in blocked_domains):
                    continue
                if ".edu.cn" not in url:
                    continue
                if url not in candidate_urls:
                    candidate_urls.append(url)

    best_url = ""
    best_score = -999

    for url in candidate_urls[:10]:
        try:
            html = fetch_external_html(session, url)

            # 🚨 核心：过滤非本校
            if not is_same_university(row, html):
                print("跳过非本校:", url)
                continue

            text = html.lower()
            score = 0

            if "招生" in text:
                score += 5
            if "招生简章" in text:
                score += 8
            if "联系方式" in text:
                score += 3
            if "合作办学" in text:
                score += 3

            if score > best_score:
                best_score = score
                best_url = url

        except Exception as e:
            print("解析失败:", url, e)

    return best_url if best_score >= 10 else ""


def search_contact_page_debug(row: dict) -> None:
    from ddgs import DDGS

    queries = build_search_queries(row)
    with DDGS() as ddgs:
        for query in queries[:3]:
            print("搜索词:", query)
            try:
                results = ddgs.text(query, max_results=5)
            except Exception as e:
                print(f"搜索失败: {query} -> {e}")
                continue
            for i, r in enumerate(results, 1):
                print(f"{i}. 标题: {r.get('title')}")
                print(f"   链接: {r.get('href') or r.get('url')}")
                print(f"   摘要: {r.get('body')}")
                print("-" * 60)


# ========= 联系方式解析 =========
def extract_contact_fields_from_html(html: str) -> dict:
    soup = BeautifulSoup(html, "html.parser")
    text = soup.get_text("\n", strip=True)

    result = {
        "contact_person": "",
        "phone": "",
        "email": "",
        "wechat": "",
        "wechat_official_account": "",
        "address": "",
        "study_mode": "",
    }

    patterns = {
        "contact_person": [
            r"联系人\s*[：:]\s*([^\n]+)",
        ],
        "phone": [
            r"(?:联系电话|咨询电话|电话|Tel|TEL|招生咨询电话|招生热线)[：:\s]*([0-9\-\(\)\s]{7,25})",
            r"(\d{3,4}-\d{7,8})",
            r"(\d{3,4}\s?\d{7,8})",
            r"(1\d{10})",
        ],
        "email": [
            r"电子邮箱\s*[：:]\s*([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})",
            r"邮箱\s*[：:]\s*([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})",
            r"([A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,})",
        ],
        "wechat": [
            r"微信号\s*[：:]\s*([^\n]+)",
            r"微信\s*[：:]\s*([^\n]+)",
        ],
        "wechat_official_account": [
            r"微信公众号\s*[：:]\s*([^\n]+)",
            r"公众号\s*[：:]\s*([^\n]+)",
        ],
        "address": [
            r"地址\s*[：:]\s*([^\n]+)",
        ],
    }

    for field, field_patterns in patterns.items():
        for pattern in field_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                if field == "phone":
                    value = re.sub(r"\s+", "", match.group(1).strip())
                    result[field] = value
                else:
                    result[field] = match.group(1).strip()
                break

    normalized_text = (
        text.replace("＋", "+")
            .replace(" ", "")
            .replace("　", "")
    )

    mode_match = re.search(r"([1-4]\+[0-4])", normalized_text)
    if mode_match:
        result["study_mode"] = mode_match.group(1)
    else:
        if (
            "四年均在国内" in text
            or "全程在国内" in text
            or "全部课程在国内完成" in text
            or "四年均在校内完成" in text
        ):
            result["study_mode"] = "4+0(推断)"

        elif (
            "两年国内两年国外" in text
            or "2年国内2年国外" in text
            or "前两年在国内后两年在国外" in text
            or ("前两年" in text and ("国外" in text or "境外" in text))
        ):
            result["study_mode"] = "2+2(推断)"

        elif (
            "三年国内一年国外" in text
            or "3年国内1年国外" in text
            or "前三年在国内第四年出国" in text
            or "前三年在国内第四年赴国外学习" in text
            or ("前三年" in text and ("国外" in text or "境外" in text or "赴英" in text or "出国" in text))
            or ("第四年" in text and ("国外" in text or "境外" in text or "赴英" in text or "出国" in text))
        ):
            result["study_mode"] = "3+1(推断)"

        elif (
            "一年国内三年国外" in text
            or "1年国内3年国外" in text
            or ("第一年" in text and ("国内" in text or "校内" in text) and "后三年" in text and ("国外" in text or "境外" in text))
        ):
            result["study_mode"] = "1+3(推断)"

    return result

def is_same_university(row: dict, html: str) -> bool:
    university = str(row.get("university_name", "")).strip()
    if not university:
        return True

    text = BeautifulSoup(html, "html.parser").get_text(" ", strip=True)
    text = re.sub(r"\s+", " ", text)

    return university in text or university[:4] in text

def auto_enrich_contacts_better(detail_records: List[Dict], session: requests.Session) -> List[Dict]:
    enriched = []
    records_to_process = detail_records[:MAX_ENRICH_RECORDS] if MAX_ENRICH_RECORDS else detail_records

    for i, row in enumerate(records_to_process, 1):
        print(f"自动搜索 {i}/{len(records_to_process)}: {row['name']}")
        new_row = row.copy()

        contact_url = search_contact_page_better(row, session)
        new_row["contact_url"] = contact_url
        new_row["contact_person"] = ""
        new_row["phone"] = ""
        new_row["email"] = ""
        new_row["wechat"] = ""
        new_row["wechat_official_account"] = ""
        new_row["address"] = ""
        new_row["study_mode"] = ""

        if contact_url:
            try:
                html = fetch_external_html(session, contact_url)
                contact_data = extract_contact_fields_from_html(html)
                print("URL:", contact_url)
                print("提取结果:", contact_data)
                new_row.update(contact_data)
            except Exception as e:
                print(f"联系方式解析失败: {contact_url} -> {e}")

        enriched.append(new_row)

    return enriched


# ========= Excel 格式 =========
def format_sheet(ws, df: pd.DataFrame):
    ws.freeze_panes = "A2"

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    col_index = {col_name: idx + 1 for idx, col_name in enumerate(df.columns)}

    def set_width(col_name: str, width: int):
        if col_name in col_index:
            col_letter = ws.cell(row=1, column=col_index[col_name]).column_letter
            ws.column_dimensions[col_letter].width = width

    set_width("country", 10)
    set_width("region", 10)
    set_width("category", 16)
    set_width("university_name", 22)
    set_width("name", 46)
    set_width("link", 16)
    set_width("level", 10)
    set_width("duration", 10)
    set_width("major_or_course", 20)
    set_width("contact_url", 40)
    set_width("contact_person", 18)
    set_width("phone", 18)
    set_width("email", 26)
    set_width("wechat", 22)
    set_width("wechat_official_account", 24)
    set_width("address", 32)

    left_align_cols = [
        "name",
        "contact_url",
        "contact_person",
        "phone",
        "email",
        "wechat",
        "wechat_official_account",
        "address",
    ]

    for col_name in left_align_cols:
        if col_name in col_index:
            col_letter = ws.cell(row=1, column=col_index[col_name]).column_letter
            for row_num in range(2, ws.max_row + 1):
                ws[f"{col_letter}{row_num}"].alignment = Alignment(
                    vertical="center", horizontal="left", wrap_text=True
                )

    if "link" in col_index:
        col_letter = ws.cell(row=1, column=col_index["link"]).column_letter
        for row_num in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row_num}"]
            url = cell.value
            if url and isinstance(url, str):
                cell.value = "查看详情"
                cell.hyperlink = url
                cell.style = "Hyperlink"
                cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)

    if "contact_url" in col_index:
        col_letter = ws.cell(row=1, column=col_index["contact_url"]).column_letter
        for row_num in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row_num}"]
            url = cell.value
            if url and isinstance(url, str) and url.startswith("http"):
                cell.hyperlink = url
                cell.style = "Hyperlink"

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 28

    country_col = col_index.get("country")
    region_col = col_index.get("region")
    category_col = col_index.get("category")

    if country_col:
        start = 2
        for _, group in df.groupby("country", sort=False):
            end = start + len(group) - 1
            if len(group) > 1:
                ws.merge_cells(start_row=start, start_column=country_col, end_row=end, end_column=country_col)
            start = end + 1

    if country_col and region_col:
        start = 2
        for _, group in df.groupby(["country", "region"], sort=False):
            end = start + len(group) - 1
            if len(group) > 1:
                ws.merge_cells(start_row=start, start_column=region_col, end_row=end, end_column=region_col)
            start = end + 1

    if country_col and region_col and category_col:
        start = 2
        for _, group in df.groupby(["country", "region", "category"], sort=False):
            end = start + len(group) - 1
            if len(group) > 1:
                ws.merge_cells(start_row=start, start_column=category_col, end_row=end, end_column=category_col)
            start = end + 1


# ========= 一个 Excel，多 tab 导出 =========
def save_all_to_excel(all_records: List[Dict], detail_records: List[Dict], filename: str = "output/crs_full_data.xlsx") -> str:
    df_all = pd.DataFrame(all_records)
    if df_all.empty:
        print("没有一级数据可导出。")
        return filename

    df_all["country_pinyin"] = df_all["country"].apply(to_pinyin)
    df_all["region_pinyin"] = df_all["region"].apply(to_pinyin)
    df_all["category_pinyin"] = df_all["category"].apply(to_pinyin)
    df_all["university_pinyin"] = df_all["university_name"].apply(to_pinyin)
    df_all["name_pinyin"] = df_all["name"].apply(to_pinyin)

    df_all = df_all.sort_values(
        by=["country_pinyin", "region_pinyin", "category_pinyin", "university_pinyin", "name_pinyin"]
    ).reset_index(drop=True)

    df_all = df_all.drop(columns=["country_pinyin", "region_pinyin", "category_pinyin", "university_pinyin", "name_pinyin"])
    df_all = df_all[[
        "country", "region", "category", "university_name", "is_985", "is_211", "is_985_211", "name", "link"
    ]]

    df_projects = df_all[df_all["category"] == "合作办学项目"].copy().reset_index(drop=True)
    df_non_985_211_projects = df_all[
        (df_all["category"] == "合作办学项目") & (df_all["is_985_211"] == False)
    ].copy().reset_index(drop=True)

    df_detail = pd.DataFrame(detail_records)
    if not df_detail.empty:
        df_detail["country_pinyin"] = df_detail["country"].apply(to_pinyin)
        df_detail["region_pinyin"] = df_detail["region"].apply(to_pinyin)
        df_detail["category_pinyin"] = df_detail["category"].apply(to_pinyin)
        df_detail["university_pinyin"] = df_detail["university_name"].apply(to_pinyin)
        df_detail["name_pinyin"] = df_detail["name"].apply(to_pinyin)

        df_detail = df_detail.sort_values(
            by=["country_pinyin", "region_pinyin", "category_pinyin", "university_pinyin", "name_pinyin"]
        ).reset_index(drop=True)

        df_detail = df_detail.drop(columns=["country_pinyin", "region_pinyin", "category_pinyin", "university_pinyin", "name_pinyin"])
        detail_cols = [
            "country",
            "region",
            "category",
            "university_name",
            "is_985",
            "is_211",
            "is_985_211",
            "name",
            "level",
            "duration",
            "admission_years",
            "admission_end_year",
            "is_active",
            "major_or_course",
            "degree_awarded",
            "foreign_degree_certificate",
            "link",
        ]
        existing_cols = [c for c in detail_cols if c in df_detail.columns]
        df_detail = df_detail[existing_cols]

    try:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_all.to_excel(writer, sheet_name="全部数据", index=False)
            df_projects.to_excel(writer, sheet_name="合作办学项目", index=False)
            df_non_985_211_projects.to_excel(writer, sheet_name="非985_211合作办学项目", index=False)
            if not df_detail.empty:
                df_detail.to_excel(writer, sheet_name="项目详情", index=False)
    except PermissionError:
        filename = "output/crs_full_data_new.xlsx"
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_all.to_excel(writer, sheet_name="全部数据", index=False)
            df_projects.to_excel(writer, sheet_name="合作办学项目", index=False)
            df_non_985_211_projects.to_excel(writer, sheet_name="非985_211合作办学项目", index=False)
            if not df_detail.empty:
                df_detail.to_excel(writer, sheet_name="项目详情", index=False)
        print(f"原文件被占用，已改存为: {filename}")

    wb = load_workbook(filename)
    format_sheet(wb["全部数据"], df_all)
    format_sheet(wb["合作办学项目"], df_projects)
    format_sheet(wb["非985_211合作办学项目"], df_non_985_211_projects)
    if not df_detail.empty:
        format_sheet(wb["项目详情"], df_detail)

    wb.save(filename)
    print(f"Excel 已导出: {filename}")
    return filename


# ========= 主流程 =========
def main():
    os.makedirs("output", exist_ok=True)

    session = requests.Session()
    session.get(REFERER, headers=CRS_HEADERS, timeout=30)

    all_records = []

    # 1. 抓列表页
    for country in COUNTRIES:
        print(f"正在抓取列表页: {country}")
        html = fetch_country_html(session, country)
        records = extract_records(html, country)
        print(f"{country} 抓到 {len(records)} 条一级数据")
        all_records.extend(records)

    if not all_records:
        print("没有抓到一级数据。")
        return

    # 2. 先过滤：只保留合作办学项目 + 去掉985/211
    project_records = [
        r for r in all_records
        if r["category"] == "合作办学项目"
        and not r["is_985_211"]
    ]
    print(f"过滤后合作办学项目共有 {len(project_records)} 条（已去掉985/211与办学机构）")

    # 3. 抓详情页，拿到 level / duration / major_or_course
    detail_records = []
    for i, row in enumerate(project_records, 1):
        print(f"正在抓详情页 {i}/{len(project_records)}: {row['link']}")
        try:
            detail_html = fetch_detail_html(session, row["link"])
            detail = extract_detail_fields(detail_html)
            merged = {**row, **detail}
            detail_records.append(merged)
        except Exception as e:
            print(f"抓取失败: {row['link']} -> {e}")

    print(f"成功抓到 {len(detail_records)} 条二级详情数据")

    # 4. 再过滤：去掉博士项目
    detail_records = [
        r for r in detail_records
        if "博士" not in str(r.get("level", ""))
    ]
    print(f"去掉博士后剩余 {len(detail_records)} 条")

    # 如果你以后想只保留本科，可以改成下面这段：
    # detail_records = [
    #     r for r in detail_records
    #     if r.get("level") == "本科"
    # ]

    # 5. 导出基础 Excel
    output_file = save_all_to_excel(all_records, detail_records)

    # 6. 自动补联系方式
    enriched = auto_enrich_contacts_better(detail_records, session)
    df_contacts = pd.DataFrame(enriched)

    if df_contacts.empty:
        print("没有联系方式数据可写入。")
        return

    contact_cols = [
        "country",
        "region",
        "category",
        "university_name",
        "name",
        "level",
        "duration",
        "admission_years",
        "admission_end_year",
        "is_active",
        "major_or_course",
        "degree_awarded",
        "foreign_degree_certificate",
        "study_mode",
        "link",
        "contact_url",
        "contact_person",
        "phone",
        "email",
        "wechat",
        "wechat_official_account",
        "address",

    ]

    existing_cols = [c for c in contact_cols if c in df_contacts.columns]
    df_contacts = df_contacts[existing_cols]

    with pd.ExcelWriter(output_file, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_contacts.to_excel(writer, sheet_name="自动联系方式", index=False)

    wb = load_workbook(output_file)
    format_sheet(wb["自动联系方式"], df_contacts)
    wb.save(output_file)

    print("联系方式已写入")


if __name__ == "__main__":
    main()
